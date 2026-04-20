#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Payment Order Web App
Flask + SQLite (local) / PostgreSQL (Railway production)
"""

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, abort)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3, os, datetime, hashlib, json, shutil
import threading, urllib.request, urllib.error
from functools import wraps
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, Image as RLImage)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font as XFont, Alignment, PatternFill, Border, Side
import io

# ── App setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-secret-key-in-production")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_FILE    = os.environ.get("DATABASE_URL", os.path.join(BASE_DIR, "payment.db"))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Register rupee-capable font if available
FONT_NORMAL = "Helvetica"; FONT_BOLD = "Helvetica-Bold"
for _name, _path in [
    ("DejaVu",      "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ("DejaVu-Bold", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
]:
    if os.path.exists(_path):
        try:
            pdfmetrics.registerFont(TTFont(_name, _path))
            FONT_NORMAL = "DejaVu"; FONT_BOLD = "DejaVu-Bold"
        except: pass

RUPEE = "Rs."   # works with all fonts

# ── DB helpers ─────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'user',
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS company (
        id INTEGER PRIMARY KEY DEFAULT 1,
        name TEXT DEFAULT 'My Company',
        address TEXT, city TEXT, state TEXT, pincode TEXT,
        phone TEXT, email TEXT, gst_no TEXT, pan_no TEXT,
        bank_details TEXT, letterhead_path TEXT, other_details TEXT
    );
    CREATE TABLE IF NOT EXISTS bank_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bank_name TEXT UNIQUE NOT NULL
    );
    CREATE TABLE IF NOT EXISTS ifsc_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ifsc_code TEXT UNIQUE NOT NULL,
        bank_name TEXT, branch TEXT, city TEXT
    );
    CREATE TABLE IF NOT EXISTS work_master (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        work_name TEXT UNIQUE NOT NULL,
        work_category TEXT DEFAULT 'Other'
    );
    CREATE TABLE IF NOT EXISTS our_accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        account_name TEXT NOT NULL,
        account_code TEXT NOT NULL,
        bank_name TEXT, account_number TEXT,
        ifsc_code TEXT, branch TEXT,
        account_type TEXT DEFAULT 'Current',
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS parties (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL, party_type TEXT,
        pan_no TEXT, aadhaar_no TEXT,
        phone TEXT, address TEXT, city TEXT,
        photo_path TEXT, notes TEXT, active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS party_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        party_id INTEGER, doc_name TEXT, doc_path TEXT,
        uploaded_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS party_accounts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        party_id INTEGER, beneficiary_name TEXT,
        relation TEXT DEFAULT 'Self',
        bank_name TEXT, account_number TEXT,
        ifsc_code TEXT, branch TEXT, is_default INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS payment_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ref_no TEXT UNIQUE, order_date TEXT,
        our_account_id INTEGER,
        cheque_no TEXT, cheque_date TEXT,
        total_gross REAL DEFAULT 0,
        total_advance REAL DEFAULT 0,
        total_net REAL DEFAULT 0,
        status TEXT DEFAULT 'Saved',
        remarks TEXT, created_by TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS payment_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER, party_id INTEGER,
        party_name TEXT, party_account_id INTEGER,
        beneficiary_name TEXT, bank_name TEXT,
        account_number TEXT, ifsc_code TEXT,
        branch TEXT, transfer_type TEXT,
        work_id INTEGER, work_name TEXT,
        details TEXT,
        advance_amount REAL DEFAULT 0,
        payment_amount REAL DEFAULT 0,
        net_amount REAL DEFAULT 0, note TEXT
    );
    """)
    # Defaults
    pw = generate_password_hash("admin123")
    c.execute("INSERT OR IGNORE INTO users (username,password_hash,role) VALUES ('admin',?,'admin')", (pw,))
    c.execute("INSERT OR IGNORE INTO company (id,name) VALUES (1,'My Farm Business')")
    for w in [("Vendor Payment","Payment"),("Salary","HR"),("Labour Wages","HR"),
               ("Rent","Expense"),("Purchase","Purchase"),("Commission","Payment"),
               ("Advance","Advance"),("Loan Repayment","Finance"),("Other","Other")]:
        c.execute("INSERT OR IGNORE INTO work_master (work_name,work_category) VALUES (?,?)", w)
    for b in ["State Bank of India","HDFC Bank","ICICI Bank","Punjab National Bank",
              "Bank of Baroda","Canara Bank","Axis Bank","UCO Bank","Indian Bank"]:
        c.execute("INSERT OR IGNORE INTO bank_master (bank_name) VALUES (?)", (b,))
    conn.commit(); conn.close()

# ── Auth helpers ───────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_role") != "admin":
            flash("Admin access required.", "error"); return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

def current_user():
    return {"id": session.get("user_id"), "username": session.get("username"),
            "role": session.get("user_role")}

# ── Utility ────────────────────────────────────────────────────────────────────
def fmt(a):
    try: return f"{RUPEE} {float(a):,.2f}"
    except: return f"{RUPEE} 0.00"

def get_fy():
    t = datetime.date.today()
    return f"{str(t.year)[2:]}-{str(t.year+1)[2:]}" if t.month >= 4 else f"{str(t.year-1)[2:]}-{str(t.year)[2:]}"

def next_ref_no(account_id, account_code):
    year = get_fy(); prefix = f"{account_code}/{year}/"
    conn = get_db()
    row = conn.execute("SELECT ref_no FROM payment_orders WHERE ref_no LIKE ? ORDER BY id DESC LIMIT 1",
                       (prefix+"%",)).fetchone()
    conn.close()
    last = int(row["ref_no"].split("/")[-1]) if row else 0
    return f"{account_code}/{year}/{str(last+1).zfill(3)}"

def today():
    return datetime.date.today().strftime("%d-%m-%Y")

IFSC_CACHE = {}

# ── Routes: Auth ───────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username","").strip()
        p = request.form.get("password","")
        conn = get_db()
        row = conn.execute("SELECT * FROM users WHERE username=? AND active=1", (u,)).fetchone()
        conn.close()
        if row and check_password_hash(row["password_hash"], p):
            session["user_id"] = row["id"]
            session["username"] = row["username"]
            session["user_role"] = row["role"]
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear(); return redirect(url_for("login"))

# ── Routes: Dashboard ──────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    conn = get_db()
    today_db = datetime.date.today().strftime("%Y-%m-%d")
    month_start = datetime.date.today().replace(day=1).strftime("%Y-%m-%d")
    stats = {
        "accounts":  conn.execute("SELECT COUNT(*) FROM our_accounts WHERE active=1").fetchone()[0],
        "parties":   conn.execute("SELECT COUNT(*) FROM parties WHERE active=1").fetchone()[0],
        "orders":    conn.execute("SELECT COUNT(*) FROM payment_orders").fetchone()[0],
        "today":     conn.execute("SELECT COALESCE(SUM(net_amount),0) FROM payment_items pi JOIN payment_orders po ON pi.order_id=po.id WHERE po.order_date=?", (today_db,)).fetchone()[0],
        "month":     conn.execute("SELECT COALESCE(SUM(net_amount),0) FROM payment_items pi JOIN payment_orders po ON pi.order_id=po.id WHERE po.order_date>=?", (month_start,)).fetchone()[0],
    }
    recent = conn.execute("""SELECT po.ref_no,po.order_date,oa.account_name,po.cheque_no,
        po.total_net,po.status,po.id,
        (SELECT COUNT(*) FROM payment_items WHERE order_id=po.id) as items
        FROM payment_orders po LEFT JOIN our_accounts oa ON po.our_account_id=oa.id
        ORDER BY po.id DESC LIMIT 15""").fetchall()
    conn.close()
    return render_template("dashboard.html", stats=stats, recent=recent, fmt=fmt)

# ── Routes: Company ────────────────────────────────────────────────────────────
@app.route("/company", methods=["GET","POST"])
@login_required
def company():
    conn = get_db()
    if request.method == "POST":
        lh_path = conn.execute("SELECT letterhead_path FROM company WHERE id=1").fetchone()["letterhead_path"]
        if "letterhead" in request.files:
            f = request.files["letterhead"]
            if f and f.filename:
                ext = os.path.splitext(secure_filename(f.filename))[1]
                dest = os.path.join(UPLOAD_DIR, f"letterhead{ext}")
                f.save(dest); lh_path = dest
        conn.execute("""UPDATE company SET name=?,address=?,city=?,state=?,pincode=?,
            phone=?,email=?,gst_no=?,pan_no=?,bank_details=?,other_details=?,letterhead_path=?
            WHERE id=1""", (
            request.form.get("name",""), request.form.get("address",""),
            request.form.get("city",""), request.form.get("state",""),
            request.form.get("pincode",""), request.form.get("phone",""),
            request.form.get("email",""), request.form.get("gst_no",""),
            request.form.get("pan_no",""), request.form.get("bank_details",""),
            request.form.get("other_details",""), lh_path))
        conn.commit(); conn.close()
        flash("Company settings saved!", "success")
        return redirect(url_for("company"))
    co = conn.execute("SELECT * FROM company WHERE id=1").fetchone()
    conn.close()
    return render_template("company.html", co=co)

# ── Routes: Our Accounts ───────────────────────────────────────────────────────
@app.route("/accounts")
@login_required
def accounts():
    conn = get_db()
    rows = conn.execute("SELECT * FROM our_accounts WHERE active=1").fetchall()
    fy = get_fy()
    conn.close()
    return render_template("accounts.html", accounts=rows, fy=fy)

@app.route("/accounts/save", methods=["POST"])
@login_required
def account_save():
    conn = get_db()
    eid = request.form.get("id")
    d = {k: request.form.get(k,"").strip() for k in
         ["account_name","account_code","bank_name","account_number","ifsc_code","branch","account_type"]}
    d["account_code"] = d["account_code"].upper()
    if eid:
        conn.execute("UPDATE our_accounts SET account_name=?,account_code=?,bank_name=?,account_number=?,ifsc_code=?,branch=?,account_type=? WHERE id=?",
            (*d.values(), eid))
    else:
        conn.execute("INSERT INTO our_accounts (account_name,account_code,bank_name,account_number,ifsc_code,branch,account_type) VALUES (?,?,?,?,?,?,?)",
            tuple(d.values()))
    conn.commit(); conn.close()
    flash("Account saved.", "success")
    return redirect(url_for("accounts"))

@app.route("/accounts/delete/<int:eid>", methods=["POST"])
@login_required
def account_delete(eid):
    conn = get_db()
    conn.execute("UPDATE our_accounts SET active=0 WHERE id=?", (eid,))
    conn.commit(); conn.close()
    flash("Account deleted.", "success")
    return redirect(url_for("accounts"))

# ── Routes: Masters ────────────────────────────────────────────────────────────
@app.route("/masters")
@login_required
def masters():
    conn = get_db()
    banks = conn.execute("SELECT * FROM bank_master ORDER BY bank_name").fetchall()
    ifscs = conn.execute("SELECT * FROM ifsc_master ORDER BY bank_name,branch").fetchall()
    works = conn.execute("SELECT * FROM work_master ORDER BY work_category,work_name").fetchall()
    conn.close()
    return render_template("masters.html", banks=banks, ifscs=ifscs, works=works)

@app.route("/masters/bank/save", methods=["POST"])
@login_required
def bank_save():
    name = request.form.get("bank_name","").strip()
    if name:
        conn = get_db()
        conn.execute("INSERT OR IGNORE INTO bank_master (bank_name) VALUES (?)", (name,))
        conn.commit(); conn.close()
        flash(f"Bank '{name}' added.", "success")
    return redirect(url_for("masters"))

@app.route("/masters/bank/delete/<int:eid>", methods=["POST"])
@login_required
def bank_delete(eid):
    conn = get_db()
    conn.execute("DELETE FROM bank_master WHERE id=?", (eid,))
    conn.commit(); conn.close()
    return redirect(url_for("masters"))

@app.route("/masters/ifsc/save", methods=["POST"])
@login_required
def ifsc_save():
    conn = get_db()
    eid = request.form.get("id")
    d = {k: request.form.get(k,"").strip() for k in ["ifsc_code","bank_name","branch","city"]}
    d["ifsc_code"] = d["ifsc_code"].upper()
    if eid:
        conn.execute("UPDATE ifsc_master SET ifsc_code=?,bank_name=?,branch=?,city=? WHERE id=?",
            (*d.values(), eid))
    else:
        conn.execute("INSERT OR REPLACE INTO ifsc_master (ifsc_code,bank_name,branch,city) VALUES (?,?,?,?)",
            tuple(d.values()))
        if d["bank_name"]:
            conn.execute("INSERT OR IGNORE INTO bank_master (bank_name) VALUES (?)", (d["bank_name"],))
    conn.commit(); conn.close()
    flash("IFSC saved.", "success")
    return redirect(url_for("masters"))

@app.route("/masters/ifsc/delete/<int:eid>", methods=["POST"])
@login_required
def ifsc_delete(eid):
    conn = get_db()
    conn.execute("DELETE FROM ifsc_master WHERE id=?", (eid,))
    conn.commit(); conn.close()
    return redirect(url_for("masters"))

@app.route("/masters/work/save", methods=["POST"])
@login_required
def work_save():
    conn = get_db()
    eid = request.form.get("id")
    name = request.form.get("work_name","").strip()
    cat  = request.form.get("work_category","Other").strip()
    if eid:
        conn.execute("UPDATE work_master SET work_name=?,work_category=? WHERE id=?", (name,cat,eid))
    else:
        conn.execute("INSERT OR IGNORE INTO work_master (work_name,work_category) VALUES (?,?)", (name,cat))
    conn.commit(); conn.close()
    flash("Work saved.", "success")
    return redirect(url_for("masters"))

@app.route("/masters/work/delete/<int:eid>", methods=["POST"])
@login_required
def work_delete(eid):
    conn = get_db()
    conn.execute("DELETE FROM work_master WHERE id=?", (eid,))
    conn.commit(); conn.close()
    return redirect(url_for("masters"))

# ── Routes: Parties ────────────────────────────────────────────────────────────
@app.route("/parties")
@login_required
def parties():
    conn = get_db()
    q = request.args.get("q",""); ptype = request.args.get("type","All")
    sql = "SELECT p.*,(SELECT COUNT(*) FROM party_accounts WHERE party_id=p.id) as acc_count FROM parties p WHERE p.active=1"
    params = []
    if q: sql += " AND p.name LIKE ?"; params.append(f"%{q}%")
    if ptype != "All": sql += " AND p.party_type=?"; params.append(ptype)
    sql += " ORDER BY p.name"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return render_template("parties.html", parties=rows, q=q, ptype=ptype)

@app.route("/parties/add", methods=["GET","POST"])
@app.route("/parties/edit/<int:pid>", methods=["GET","POST"])
@login_required
def party_form(pid=None):
    conn = get_db()
    if request.method == "POST":
        photo_path = None
        if pid:
            existing = conn.execute("SELECT photo_path FROM parties WHERE id=?", (pid,)).fetchone()
            photo_path = existing["photo_path"] if existing else None
        if "photo" in request.files:
            f = request.files["photo"]
            if f and f.filename:
                ext = os.path.splitext(secure_filename(f.filename))[1]
                dest = os.path.join(UPLOAD_DIR, f"party_{pid or 'new'}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}{ext}")
                f.save(dest); photo_path = dest
        d = {k: request.form.get(k,"").strip() for k in
             ["name","party_type","pan_no","aadhaar_no","phone","address","city","notes"]}
        if pid:
            conn.execute("UPDATE parties SET name=?,party_type=?,pan_no=?,aadhaar_no=?,phone=?,address=?,city=?,notes=?,photo_path=? WHERE id=?",
                (*d.values(), photo_path, pid))
            conn.commit(); conn.close()
            flash("Party updated.", "success")
            return redirect(url_for("party_detail", pid=pid))
        else:
            cur = conn.execute("INSERT INTO parties (name,party_type,pan_no,aadhaar_no,phone,address,city,notes,photo_path) VALUES (?,?,?,?,?,?,?,?,?)",
                (*d.values(), photo_path))
            new_pid = cur.lastrowid
            conn.commit(); conn.close()
            flash("Party added!", "success")
            return redirect(url_for("party_detail", pid=new_pid))
    party = conn.execute("SELECT * FROM parties WHERE id=?", (pid,)).fetchone() if pid else None
    conn.close()
    return render_template("party_form.html", party=party)

@app.route("/parties/<int:pid>")
@login_required
def party_detail(pid):
    conn = get_db()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (pid,)).fetchone()
    if not party: abort(404)
    accounts = conn.execute("SELECT * FROM party_accounts WHERE party_id=? ORDER BY is_default DESC", (pid,)).fetchall()
    docs = conn.execute("SELECT * FROM party_documents WHERE party_id=? ORDER BY id DESC", (pid,)).fetchall()
    banks = conn.execute("SELECT bank_name FROM bank_master ORDER BY bank_name").fetchall()
    conn.close()
    return render_template("party_detail.html", party=party, accounts=accounts, docs=docs, banks=banks)

@app.route("/parties/<int:pid>/delete", methods=["POST"])
@login_required
def party_delete(pid):
    conn = get_db()
    conn.execute("UPDATE parties SET active=0 WHERE id=?", (pid,))
    conn.commit(); conn.close()
    flash("Party deleted.", "success")
    return redirect(url_for("parties"))

@app.route("/parties/<int:pid>/accounts/save", methods=["POST"])
@login_required
def party_account_save(pid):
    conn = get_db()
    eid = request.form.get("id")
    d = {k: request.form.get(k,"").strip() for k in
         ["beneficiary_name","relation","bank_name","account_number","ifsc_code","branch"]}
    d["ifsc_code"] = d["ifsc_code"].upper()
    is_def = 1 if request.form.get("is_default") else 0
    if is_def:
        conn.execute("UPDATE party_accounts SET is_default=0 WHERE party_id=?", (pid,))
    if eid:
        conn.execute("UPDATE party_accounts SET beneficiary_name=?,relation=?,bank_name=?,account_number=?,ifsc_code=?,branch=?,is_default=? WHERE id=?",
            (*d.values(), is_def, eid))
    else:
        conn.execute("INSERT INTO party_accounts (party_id,beneficiary_name,relation,bank_name,account_number,ifsc_code,branch,is_default) VALUES (?,?,?,?,?,?,?,?)",
            (pid, *d.values(), is_def))
    conn.commit(); conn.close()
    flash("Account saved.", "success")
    return redirect(url_for("party_detail", pid=pid))

@app.route("/parties/<int:pid>/accounts/<int:aid>/delete", methods=["POST"])
@login_required
def party_account_delete(pid, aid):
    conn = get_db()
    conn.execute("DELETE FROM party_accounts WHERE id=? AND party_id=?", (aid, pid))
    conn.commit(); conn.close()
    flash("Account deleted.", "success")
    return redirect(url_for("party_detail", pid=pid))

@app.route("/parties/<int:pid>/accounts/<int:aid>/default", methods=["POST"])
@login_required
def party_account_set_default(pid, aid):
    conn = get_db()
    conn.execute("UPDATE party_accounts SET is_default=0 WHERE party_id=?", (pid,))
    conn.execute("UPDATE party_accounts SET is_default=1 WHERE id=?", (aid,))
    conn.commit(); conn.close()
    return redirect(url_for("party_detail", pid=pid))

@app.route("/parties/<int:pid>/documents/upload", methods=["POST"])
@login_required
def doc_upload(pid):
    f = request.files.get("doc_file")
    dname = request.form.get("doc_name","").strip()
    if f and f.filename and dname:
        ext = os.path.splitext(secure_filename(f.filename))[1]
        dest = os.path.join(UPLOAD_DIR, f"doc_{pid}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}{ext}")
        f.save(dest)
        conn = get_db()
        conn.execute("INSERT INTO party_documents (party_id,doc_name,doc_path) VALUES (?,?,?)",
                     (pid, dname, dest))
        conn.commit(); conn.close()
        flash("Document uploaded.", "success")
    return redirect(url_for("party_detail", pid=pid))

@app.route("/documents/<int:did>/download")
@login_required
def doc_download(did):
    conn = get_db()
    doc = conn.execute("SELECT * FROM party_documents WHERE id=?", (did,)).fetchone()
    conn.close()
    if doc and os.path.exists(doc["doc_path"]):
        return send_file(doc["doc_path"], as_attachment=False,
                         download_name=doc["doc_name"] + os.path.splitext(doc["doc_path"])[1])
    abort(404)

@app.route("/parties/<int:pid>/documents/<int:did>/delete", methods=["POST"])
@login_required
def doc_delete(pid, did):
    conn = get_db()
    doc = conn.execute("SELECT doc_path FROM party_documents WHERE id=?", (did,)).fetchone()
    conn.execute("DELETE FROM party_documents WHERE id=?", (did,))
    conn.commit(); conn.close()
    if doc and doc["doc_path"] and os.path.exists(doc["doc_path"]):
        try: os.remove(doc["doc_path"])
        except: pass
    flash("Document deleted.", "success")
    return redirect(url_for("party_detail", pid=pid))

# ── Routes: Payment Orders ─────────────────────────────────────────────────────
@app.route("/orders/new")
@login_required
def order_new():
    conn = get_db()
    accs    = conn.execute("SELECT * FROM our_accounts WHERE active=1").fetchall()
    parties = conn.execute("SELECT id,name FROM parties WHERE active=1 ORDER BY name").fetchall()
    works   = conn.execute("SELECT * FROM work_master ORDER BY work_name").fetchall()
    conn.close()
    return render_template("order_form.html", accs=accs, parties=parties,
                           works=works, today=today(), fy=get_fy(), order=None)

@app.route("/orders/edit/<int:oid>")
@login_required
def order_edit(oid):
    conn = get_db()
    order = conn.execute("SELECT po.*,oa.account_name,oa.account_code FROM payment_orders po LEFT JOIN our_accounts oa ON po.our_account_id=oa.id WHERE po.id=?", (oid,)).fetchone()
    items = conn.execute("SELECT * FROM payment_items WHERE order_id=?", (oid,)).fetchall()
    accs    = conn.execute("SELECT * FROM our_accounts WHERE active=1").fetchall()
    parties = conn.execute("SELECT id,name FROM parties WHERE active=1 ORDER BY name").fetchall()
    works   = conn.execute("SELECT * FROM work_master ORDER BY work_name").fetchall()
    conn.close()
    return render_template("order_form.html", accs=accs, parties=parties,
                           works=works, today=today(), fy=get_fy(),
                           order=order, items=items)

@app.route("/orders/save", methods=["POST"])
@login_required
def order_save():
    conn = get_db()
    oid = request.form.get("order_id") or None
    acc_id = request.form.get("our_account_id")
    acc = conn.execute("SELECT * FROM our_accounts WHERE id=?", (acc_id,)).fetchone()
    if not acc:
        flash("Please select an account.", "error")
        return redirect(url_for("order_new"))

    date_raw = request.form.get("order_date","")
    try: date_db = datetime.datetime.strptime(date_raw, "%d-%m-%Y").strftime("%Y-%m-%d")
    except: date_db = datetime.date.today().strftime("%Y-%m-%d")

    # Collect items from JSON
    items_json = request.form.get("items_json","[]")
    try: items = json.loads(items_json)
    except: items = []

    total_gross = sum(float(i.get("payment_amount",0)) for i in items)
    total_adv   = sum(float(i.get("advance_amount",0)) for i in items)
    total_net   = sum(float(i.get("net_amount",0)) for i in items)

    if oid:
        ref_no = request.form.get("ref_no")
        conn.execute("UPDATE payment_orders SET order_date=?,our_account_id=?,cheque_no=?,cheque_date=?,total_gross=?,total_advance=?,total_net=?,status='Saved' WHERE id=?",
            (date_db, acc_id, request.form.get("cheque_no",""),
             request.form.get("cheque_date",""), total_gross, total_adv, total_net, oid))
        conn.execute("DELETE FROM payment_items WHERE order_id=?", (oid,))
    else:
        ref_no = next_ref_no(acc["id"], acc["account_code"])
        cur = conn.execute("INSERT INTO payment_orders (ref_no,order_date,our_account_id,cheque_no,cheque_date,total_gross,total_advance,total_net,status,created_by) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (ref_no, date_db, acc_id, request.form.get("cheque_no",""),
             request.form.get("cheque_date",""), total_gross, total_adv, total_net, "Saved",
             session.get("username")))
        oid = cur.lastrowid

    for item in items:
        conn.execute("INSERT INTO payment_items (order_id,party_id,party_name,party_account_id,beneficiary_name,bank_name,account_number,ifsc_code,branch,transfer_type,work_id,work_name,details,advance_amount,payment_amount,net_amount,note) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (oid, item.get("party_id"), item.get("party_name"),
             item.get("party_account_id",0), item.get("beneficiary_name",""),
             item.get("bank_name",""), item.get("account_number",""),
             item.get("ifsc_code",""), item.get("branch",""),
             item.get("transfer_type","NEFT"), item.get("work_id"),
             item.get("work_name",""), item.get("details",""),
             float(item.get("advance_amount",0)), float(item.get("payment_amount",0)),
             float(item.get("net_amount",0)), item.get("note","")))
    conn.commit(); conn.close()
    flash(f"Order {ref_no} saved! Net Total: {fmt(total_net)}", "success")
    return redirect(url_for("order_detail", oid=oid))

@app.route("/orders/<int:oid>")
@login_required
def order_detail(oid):
    conn = get_db()
    order = conn.execute("SELECT po.*,oa.account_name,oa.account_code,oa.bank_name as our_bank,oa.account_number as our_acc_no,oa.ifsc_code as our_ifsc,oa.branch as our_branch FROM payment_orders po LEFT JOIN our_accounts oa ON po.our_account_id=oa.id WHERE po.id=?", (oid,)).fetchone()
    items = conn.execute("SELECT * FROM payment_items WHERE order_id=?", (oid,)).fetchall()
    conn.close()
    if not order: abort(404)
    return render_template("order_detail.html", order=order, items=items, fmt=fmt)

@app.route("/orders/<int:oid>/delete", methods=["POST"])
@login_required
def order_delete(oid):
    conn = get_db()
    conn.execute("DELETE FROM payment_items WHERE order_id=?", (oid,))
    conn.execute("DELETE FROM payment_orders WHERE id=?", (oid,))
    conn.commit(); conn.close()
    flash("Order deleted.", "success")
    return redirect(url_for("history"))

@app.route("/history")
@login_required
def history():
    conn = get_db()
    from_d = request.args.get("from","01-04-"+str(datetime.date.today().year if datetime.date.today().month>=4 else datetime.date.today().year-1))
    to_d   = request.args.get("to", today())
    party_f= request.args.get("party","")
    acc_f  = request.args.get("account","")
    try:
        fd = datetime.datetime.strptime(from_d,"%d-%m-%Y").strftime("%Y-%m-%d")
        td = datetime.datetime.strptime(to_d,"%d-%m-%Y").strftime("%Y-%m-%d")
    except: fd="2000-01-01"; td="2099-12-31"

    sql = """SELECT po.id,po.ref_no,po.order_date,oa.account_name,po.cheque_no,
        po.total_gross,po.total_advance,po.total_net,po.status,po.created_by,
        (SELECT COUNT(*) FROM payment_items WHERE order_id=po.id) as items
        FROM payment_orders po LEFT JOIN our_accounts oa ON po.our_account_id=oa.id
        WHERE po.order_date BETWEEN ? AND ?"""
    params = [fd, td]
    if party_f:
        sql += " AND po.id IN (SELECT DISTINCT order_id FROM payment_items WHERE party_name LIKE ?)"
        params.append(f"%{party_f}%")
    if acc_f:
        sql += " AND oa.account_name=?"; params.append(acc_f)
    sql += " ORDER BY po.order_date DESC, po.id DESC"

    orders = conn.execute(sql, params).fetchall()
    acc_list = conn.execute("SELECT DISTINCT account_name FROM our_accounts WHERE active=1").fetchall()
    total = sum(o["total_net"] for o in orders)
    conn.close()
    return render_template("history.html", orders=orders, from_d=from_d, to_d=to_d,
                           party_f=party_f, acc_f=acc_f, acc_list=acc_list, total=total, fmt=fmt)

# ── Routes: PDF ────────────────────────────────────────────────────────────────
@app.route("/orders/<int:oid>/pdf")
@login_required
def order_pdf(oid):
    conn = get_db()
    order = conn.execute("SELECT po.*,oa.* FROM payment_orders po LEFT JOIN our_accounts oa ON po.our_account_id=oa.id WHERE po.id=?", (oid,)).fetchone()
    items = conn.execute("SELECT * FROM payment_items WHERE order_id=?", (oid,)).fetchall()
    company = conn.execute("SELECT * FROM company WHERE id=1").fetchone()
    conn.close()
    buf = io.BytesIO()
    build_bank_letter_pdf(buf, list(items), order, company)
    buf.seek(0)
    fname = f"{order['ref_no'].replace('/','_')}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=False, download_name=fname)

# ── Routes: Excel ──────────────────────────────────────────────────────────────
@app.route("/orders/<int:oid>/excel")
@login_required
def order_excel(oid):
    conn = get_db()
    order = conn.execute("SELECT po.*,oa.account_name FROM payment_orders po LEFT JOIN our_accounts oa ON po.our_account_id=oa.id WHERE po.id=?", (oid,)).fetchone()
    items = conn.execute("SELECT * FROM payment_items WHERE order_id=?", (oid,)).fetchall()
    conn.close()
    buf = io.BytesIO()
    build_excel(buf, list(items), order)
    buf.seek(0)
    fname = f"{order['ref_no'].replace('/','_')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)

# ── Routes: Reports ────────────────────────────────────────────────────────────
@app.route("/reports")
@login_required
def reports():
    conn = get_db()
    fy_start = "01-04-" + str(datetime.date.today().year if datetime.date.today().month>=4 else datetime.date.today().year-1)
    from_d = request.args.get("from", fy_start)
    to_d   = request.args.get("to", today())
    party_f= request.args.get("party","All")
    work_f = request.args.get("work","All")
    ttype_f= request.args.get("ttype","All")
    acc_f  = request.args.get("account","All")
    group_f= request.args.get("group","None")
    try:
        fd = datetime.datetime.strptime(from_d,"%d-%m-%Y").strftime("%Y-%m-%d")
        td = datetime.datetime.strptime(to_d,"%d-%m-%Y").strftime("%Y-%m-%d")
    except: fd="2000-01-01"; td="2099-12-31"

    sql = """SELECT po.ref_no,po.order_date,oa.account_name,pi.party_name,
        pi.beneficiary_name,pi.bank_name,pi.transfer_type,pi.work_name,
        pi.payment_amount,pi.advance_amount,pi.net_amount
        FROM payment_items pi
        JOIN payment_orders po ON pi.order_id=po.id
        LEFT JOIN our_accounts oa ON po.our_account_id=oa.id
        WHERE po.order_date BETWEEN ? AND ?"""
    params = [fd, td]
    if party_f != "All": sql += " AND pi.party_name=?"; params.append(party_f)
    if work_f  != "All": sql += " AND pi.work_name=?";  params.append(work_f)
    if ttype_f != "All": sql += " AND pi.transfer_type=?"; params.append(ttype_f)
    if acc_f   != "All": sql += " AND oa.account_name=?"; params.append(acc_f)
    sql += " ORDER BY po.order_date DESC"

    rows = conn.execute(sql, params).fetchall()
    parties_list = conn.execute("SELECT DISTINCT name FROM parties WHERE active=1 ORDER BY name").fetchall()
    works_list   = conn.execute("SELECT DISTINCT work_name FROM work_master ORDER BY work_name").fetchall()
    acc_list     = conn.execute("SELECT DISTINCT account_name FROM our_accounts WHERE active=1").fetchall()
    conn.close()

    total = {"gross": sum(r["payment_amount"] for r in rows),
             "adv":   sum(r["advance_amount"] for r in rows),
             "net":   sum(r["net_amount"] for r in rows)}
    return render_template("reports.html", rows=rows, total=total, fmt=fmt,
                           from_d=from_d, to_d=to_d,
                           party_f=party_f, work_f=work_f, ttype_f=ttype_f,
                           acc_f=acc_f, group_f=group_f,
                           parties_list=parties_list, works_list=works_list,
                           acc_list=acc_list)

# ── Routes: Users ──────────────────────────────────────────────────────────────
@app.route("/users")
@login_required
@admin_required
def users():
    conn = get_db()
    rows = conn.execute("SELECT id,username,role,active FROM users").fetchall()
    conn.close()
    return render_template("users.html", users=rows)

@app.route("/users/save", methods=["POST"])
@login_required
@admin_required
def user_save():
    conn = get_db()
    eid = request.form.get("id")
    username = request.form.get("username","").strip()
    password = request.form.get("password","").strip()
    role     = request.form.get("role","user")
    if eid:
        if password:
            conn.execute("UPDATE users SET role=?,password_hash=? WHERE id=?",
                         (role, generate_password_hash(password), eid))
        else:
            conn.execute("UPDATE users SET role=? WHERE id=?", (role, eid))
    else:
        if not username or not password:
            flash("Username and password required.", "error")
            conn.close(); return redirect(url_for("users"))
        try:
            conn.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)",
                         (username, generate_password_hash(password), role))
        except:
            flash("Username already exists.", "error")
            conn.close(); return redirect(url_for("users"))
    conn.commit(); conn.close()
    flash("User saved.", "success")
    return redirect(url_for("users"))

@app.route("/users/<int:uid>/toggle", methods=["POST"])
@login_required
@admin_required
def user_toggle(uid):
    if uid == session["user_id"]:
        flash("Cannot deactivate yourself.", "error")
        return redirect(url_for("users"))
    conn = get_db()
    u = conn.execute("SELECT active FROM users WHERE id=?", (uid,)).fetchone()
    conn.execute("UPDATE users SET active=? WHERE id=?", (0 if u["active"] else 1, uid))
    conn.commit(); conn.close()
    return redirect(url_for("users"))

# ── Routes: API (for AJAX) ─────────────────────────────────────────────────────
@app.route("/api/ifsc/<code>")
@login_required
def api_ifsc(code):
    code = code.strip().upper()
    conn = get_db()
    local = conn.execute("SELECT * FROM ifsc_master WHERE ifsc_code=?", (code,)).fetchone()
    conn.close()
    if local:
        return jsonify({"source":"local","bank":local["bank_name"],"branch":local["branch"],"city":local["city"],"ifsc":code})
    # Try online
    if code in IFSC_CACHE:
        return jsonify(IFSC_CACHE[code])
    try:
        req = urllib.request.Request(f"https://ifsc.razorpay.com/{code}",
                                     headers={"User-Agent":"PaymentOrder/1.0"})
        with urllib.request.urlopen(req, timeout=5) as r:
            data = json.loads(r.read())
        result = {"source":"online","bank":data.get("BANK",""),"branch":data.get("BRANCH",""),
                  "city":data.get("CITY",""),"state":data.get("STATE",""),"ifsc":code}
        IFSC_CACHE[code] = result
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO ifsc_master (ifsc_code,bank_name,branch,city) VALUES (?,?,?,?)",
                     (code, result["bank"], result["branch"], result["city"]))
        if result["bank"]:
            conn.execute("INSERT OR IGNORE INTO bank_master (bank_name) VALUES (?)", (result["bank"],))
        conn.commit(); conn.close()
        return jsonify(result)
    except urllib.error.HTTPError as e:
        return jsonify({"error": f"IFSC not found (HTTP {e.code})"}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/party/<int:pid>/accounts")
@login_required
def api_party_accounts(pid):
    conn = get_db()
    accs = conn.execute("SELECT * FROM party_accounts WHERE party_id=? ORDER BY is_default DESC", (pid,)).fetchall()
    conn.close()
    return jsonify([dict(a) for a in accs])

@app.route("/api/ref_no/<int:acc_id>")
@login_required
def api_ref_no(acc_id):
    conn = get_db()
    acc = conn.execute("SELECT account_code FROM our_accounts WHERE id=?", (acc_id,)).fetchone()
    conn.close()
    if not acc: return jsonify({"error":"Account not found"}), 404
    return jsonify({"ref_no": next_ref_no(acc_id, acc["account_code"])})

# ── PDF builder ────────────────────────────────────────────────────────────────
def amount_in_words(amount):
    ones=["","ONE","TWO","THREE","FOUR","FIVE","SIX","SEVEN","EIGHT","NINE","TEN","ELEVEN","TWELVE","THIRTEEN","FOURTEEN","FIFTEEN","SIXTEEN","SEVENTEEN","EIGHTEEN","NINETEEN"]
    tens=["","","TWENTY","THIRTY","FORTY","FIFTY","SIXTY","SEVENTY","EIGHTY","NINETY"]
    def td(n):
        return ones[n] if n<20 else tens[n//10]+(" "+ones[n%10] if n%10 else "")
    def thd(n):
        return ones[n//100]+" HUNDRED"+(" "+td(n%100) if n%100 else "") if n>=100 else td(n)
    try: n=int(round(float(amount)))
    except: return "ZERO ONLY"
    if n==0: return "ZERO ONLY"
    parts=[]
    cr=n//10000000; n%=10000000; lk=n//100000; n%=100000; th=n//1000; n%=1000
    if cr: parts.append(thd(cr)+" CRORE")
    if lk: parts.append(thd(lk)+" LAKH")
    if th: parts.append(thd(th)+" THOUSAND")
    if n:  parts.append(thd(n))
    return " ".join(parts)+" ONLY"

def build_bank_letter_pdf(buf, items, order, company):
    doc = SimpleDocTemplate(buf, pagesize=A4,
          topMargin=0.5*cm, bottomMargin=1.5*cm,
          leftMargin=1.8*cm, rightMargin=1.8*cm)
    W=17.4*cm; elements=[]
    FN=FONT_NORMAL; FB=FONT_BOLD

    def ps(name,**kw): return ParagraphStyle(name,fontName=kw.pop("bold",False) and FB or FN,**kw)
    normal=ps("N",fontSize=10,leading=14)
    bold_s=ps("B",fontSize=10,leading=14,bold=True)
    hdr_c =ps("HC",fontSize=9,bold=True,textColor=colors.black,alignment=TA_CENTER,leading=11)
    dat_c =ps("DC",fontSize=9,textColor=colors.black,alignment=TA_CENTER,leading=11)
    dat_r =ps("DR",fontSize=9,textColor=colors.black,alignment=TA_RIGHT,leading=11)
    dat_br=ps("DBR",fontSize=9,bold=True,textColor=colors.black,alignment=TA_RIGHT,leading=11)

    co_name   = company["name"] if company else "My Company"
    acc_no    = order["account_number"] if "account_number" in order.keys() else "N/A"
    our_bank  = order["our_bank"] if "our_bank" in order.keys() else ""
    our_branch= order["our_branch"] if "our_branch" in order.keys() else ""

    # Letterhead
    lh = company["letterhead_path"] if company else None
    if lh and os.path.exists(str(lh)):
        try: elements.append(RLImage(lh, width=W, height=3.8*cm))
        except: pass
    else:
        lh_t = Table([[Paragraph(f"<b>{co_name}</b>",
            ParagraphStyle("LH",fontName=FB,fontSize=16,alignment=TA_CENTER,
                           textColor=colors.HexColor("#1A5276")))]],
            colWidths=[W])
        lh_t.setStyle(TableStyle([("BOX",(0,0),(-1,-1),1,colors.HexColor("#1A5276")),
            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
        elements.append(lh_t)
    elements.append(Spacer(1,0.25*cm))

    # Address + Date block
    try: disp_date = datetime.datetime.strptime(str(order["order_date"]),"%Y-%m-%d").strftime("%d-%m-%Y")
    except: disp_date = str(order["order_date"])

    left_txt = f"To<br/>The Manager<br/><b>{our_bank}</b><br/>{our_branch}<br/>{company['city'] if company and company['city'] else ''}"
    right_rows=[
        [Paragraph("Date",bold_s), Paragraph(disp_date,normal)],
        [Paragraph("REF NO",bold_s), Paragraph(f"<b>{order['ref_no']}</b>",
            ParagraphStyle("RN",fontName=FB,fontSize=10,textColor=colors.HexColor("#1A5276")))],
        [Paragraph("CHEQUE NO",bold_s), Paragraph(str(order["cheque_no"] or "-"),normal)],
        [Paragraph("CHEQUE DATE",bold_s), Paragraph(str(order["cheque_date"] or "-"),normal)],
    ]
    rt = Table(right_rows, colWidths=[3.2*cm,4.5*cm])
    rt.setStyle(TableStyle([("ALIGN",(0,0),(0,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("LEFTPADDING",(0,0),(-1,-1),4),("RIGHTPADDING",(0,0),(-1,-1),4)]))
    addr_t = Table([[Paragraph(left_txt, ParagraphStyle("A",fontName=FN,fontSize=10,leading=16)),rt]],
                   colWidths=[9.5*cm,7.9*cm])
    addr_t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0)]))
    elements.append(addr_t)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("<b>Subject: Release of Payment</b>",
        ParagraphStyle("S",fontName=FB,fontSize=10,spaceAfter=6)))
    elements.append(Paragraph("Sir,",normal))
    elements.append(Spacer(1,0.1*cm))
    elements.append(Paragraph("Respected Sir,",normal))
    elements.append(Spacer(1,0.1*cm))
    elements.append(Paragraph(
        f"You are requested to kindly RTGS/NEFT the following amount from our A/C No. - <b>{acc_no}</b>",
        ParagraphStyle("B2",fontName=FN,fontSize=10,leading=14,spaceAfter=4)))

    # Company name sub-header
    co_t = Table([[Paragraph(co_name, ParagraphStyle("CSH",fontName=FB,fontSize=10,alignment=TA_CENTER))]],
                 colWidths=[W])
    co_t.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.black),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]))
    elements.append(co_t)

    # Split items A/B
    our_norm = (our_bank or "").strip().upper()
    items_a=[i for i in items if not (our_norm and our_norm in (i["bank_name"] or "").upper())]
    items_b=[i for i in items if our_norm and our_norm in (i["bank_name"] or "").upper()]
    if not items_a and not items_b: items_a = list(items)

    cw=[0.7*cm,4.2*cm,3.8*cm,3.2*cm,2.5*cm,2.8*cm,2.2*cm]
    hdr_row = [[Paragraph(h,hdr_c) for h in ["#","NAME","BENEFICIARY A/C No","BANK","BRANCH","IFSC","AMOUNT"]]]

    def item_rows(lst, start=1):
        rows=[]
        for i,it in enumerate(lst,start):
            amt=float(it["net_amount"] or 0)
            rows.append([Paragraph(str(i),dat_c),
                Paragraph(str(it["beneficiary_name"] or it["party_name"]),dat_c),
                Paragraph(str(it["account_number"] or ""),dat_c),
                Paragraph(str(it["bank_name"] or ""),dat_c),
                Paragraph(str(it["branch"] or ""),dat_c),
                Paragraph(str(it["ifsc_code"] or ""),dat_c),
                Paragraph(f"{RUPEE} {amt:,.2f}",dat_r)])
        return rows

    bdr=TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.black),
        ("INNERGRID",(0,0),(-1,-1),0.5,colors.black),
        ("BACKGROUND",(0,0),(-1,0),colors.white),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("FONTNAME",(0,0),(-1,0),FB),("FONTNAME",(0,-1),(-1,-1),FB),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#F2F2F2"))])

    serial=1
    if items_a:
        lbl_a=Table([[Paragraph("(A)  NEFT / RTGS — Other Banks",
            ParagraphStyle("LA",fontName=FB,fontSize=9,alignment=TA_CENTER))]],colWidths=[W])
        lbl_a.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.black),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EBF5FB"))]))
        elements.append(lbl_a)
        ta_sum=sum(float(i["net_amount"] or 0) for i in items_a)
        ta_data=hdr_row+item_rows(items_a,serial)
        ta_data.append(["","","","","","TOTAL (A)",Paragraph(f"{RUPEE} {ta_sum:,.2f}",dat_br)])
        serial+=len(items_a)
        ta=Table(ta_data,colWidths=cw,repeatRows=1); ta.setStyle(bdr)
        elements.append(ta)

    if items_b:
        lbl_b=Table([[Paragraph("(B) TRANSFER — Same Bank",
            ParagraphStyle("LB",fontName=FB,fontSize=9,alignment=TA_CENTER))]],colWidths=[W])
        lbl_b.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.5,colors.black),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F9F9F9"))]))
        elements.append(lbl_b)
        tb_sum=sum(float(i["net_amount"] or 0) for i in items_b)
        tb_data=hdr_row+item_rows(items_b,serial)
        tb_data.append(["","","","","","TOTAL (B)",Paragraph(f"{RUPEE} {tb_sum:,.2f}",dat_br)])
        tb=Table(tb_data,colWidths=cw,repeatRows=1); tb.setStyle(bdr)
        elements.append(tb)

    grand=sum(float(i["net_amount"] or 0) for i in items)
    gt=Table([["","","","","",
        Paragraph("GRAND TOTAL (A+B)",ParagraphStyle("GT",fontName=FB,fontSize=9,alignment=TA_CENTER)),
        Paragraph(f"{RUPEE} {grand:,.2f}",dat_br)]],colWidths=cw)
    gt.setStyle(TableStyle([("BOX",(0,0),(-1,-1),0.8,colors.black),
        ("INNERGRID",(0,0),(-1,-1),0.5,colors.black),
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EEEEEE")),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
        ("FONTNAME",(0,0),(-1,-1),FB)]))
    elements.append(gt)
    elements.append(Spacer(1,0.2*cm))
    elements.append(Paragraph(f"<b>AMOUNT: {amount_in_words(grand)}</b>",
        ParagraphStyle("AW",fontName=FB,fontSize=10,spaceAfter=6)))
    elements.append(Spacer(1,0.3*cm))
    elements.append(Paragraph("Thanking You,",normal))
    elements.append(Spacer(1,0.3*cm))
    elements.append(Paragraph("For",normal))
    elements.append(Paragraph(f"<b>{co_name}</b>",
        ParagraphStyle("SC",fontName=FB,fontSize=11,spaceAfter=4)))
    elements.append(Spacer(1,1.2*cm))
    elements.append(Paragraph("_________________________",normal))
    elements.append(Paragraph("<b>Authorised Signatory</b>",bold_s))
    doc.build(elements)

def build_excel(buf, items, order):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Payment Order"
    hf=PatternFill("solid",fgColor="1A5276"); hfont=XFont(bold=True,color="FFFFFF",size=10)
    ca=Alignment(horizontal="center",vertical="center")
    alt=PatternFill("solid",fgColor="EBF5FB")
    bd=Border(left=Side(style="thin",color="AED6F1"),right=Side(style="thin",color="AED6F1"),
              top=Side(style="thin",color="AED6F1"),bottom=Side(style="thin",color="AED6F1"))
    ws.merge_cells("A1:L1"); ws["A1"]=order["ref_no"]
    ws["A1"].font=XFont(bold=True,size=14,color="1A5276")
    ws["A1"].alignment=Alignment(horizontal="center"); ws.row_dimensions[1].height=26
    ws["A3"]="Date:"; ws["B3"]=order["order_date"]
    ws["D3"]="Cheque:"; ws["E3"]=f"{order['cheque_no'] or ''} / {order['cheque_date'] or ''}"
    ws["A4"]="Account:"; ws["B4"]=order["account_name"] if "account_name" in order.keys() else ""
    headers=["#","Party","Beneficiary","Bank","Account No","IFSC","Type","Work","Details","Amount","Advance","Net"]
    widths=[6,20,20,18,20,14,8,16,20,14,12,14]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c=ws.cell(row=6,column=i,value=h)
        c.fill=hf; c.font=hfont; c.alignment=ca; c.border=bd
        ws.column_dimensions[chr(64+i)].width=w
    total=0
    for j,item in enumerate(items,1):
        row=6+j
        vals=[j,item["party_name"],item["beneficiary_name"],item["bank_name"],
              item["account_number"],item["ifsc_code"],item["transfer_type"],
              item["work_name"],item["details"],
              float(item["payment_amount"] or 0),float(item["advance_amount"] or 0),
              float(item["net_amount"] or 0)]
        for k,val in enumerate(vals,1):
            c=ws.cell(row=row,column=k,value=val)
            c.alignment=ca; c.border=bd
            if j%2==0: c.fill=alt
        total+=float(item["net_amount"] or 0)
    tr=6+len(items)+1
    ws.cell(row=tr,column=11,value="NET TOTAL").font=XFont(bold=True,size=11)
    tc=ws.cell(row=tr,column=12,value=total); tc.font=XFont(bold=True,size=12,color="1A5276")
    tc.number_format='#,##0.00'
    wb.save(buf)

# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") == "development"
    app.run(host="0.0.0.0", port=port, debug=debug)
